from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from finance_ai.config import EXPORT_DIR

TEMPLATE_PATH = EXPORT_DIR / "finance_template.xlsx"


SHEETS = {
    "Accounts": [
        "Name",
        "Account Type",
        "Institution",
        "Current Balance",
        "Notes",
    ],
    "Categories": [
        "Name",
        "Category Type",
        "Essential",
    ],
    "Transactions": [
        "Transaction Date",
        "Merchant",
        "Description",
        "Amount",
        "Account Name",
        "Category Name",
        "Notes",
    ],
    "Debts": [
        "Name",
        "Lender",
        "Balance",
        "Interest Rate",
        "Minimum Payment",
        "Due Day",
        "Notes",
    ],
    "Assets": [
        "Name",
        "Asset Type",
        "Current Value",
        "Notes",
    ],
    "Budgets": [
        "Month",
        "Category Name",
        "Budgeted Amount",
    ],
    "Goals": [
        "Name",
        "Target Amount",
        "Current Amount",
        "Target Date",
        "Notes",
    ],
}


EXAMPLE_ROWS = {
    "Accounts": [
        ["Checking", "checking", "Bank Name", 2500, "Main household checking"],
        ["Savings", "savings", "Bank Name", 10000, "Emergency fund"],
    ],
    "Categories": [
        ["Salary", "income", ""],
        ["Groceries", "expense", "yes"],
        ["Mortgage", "expense", "yes"],
        ["Debt Payment", "expense", "yes"],
        ["Restaurants", "expense", "no"],
        ["Streaming", "expense", "no"],
    ],
    "Transactions": [
        ["2026-06-01", "Costco", "Groceries", -225.50, "Checking", "Groceries", ""],
        ["2026-06-01", "Employer", "Paycheck", 3500.00, "Checking", "Salary", ""],
    ],
    "Debts": [
        ["Truck Loan", "Credit Union", 22000, 6.49, 475, 15, ""],
        ["Credit Card", "Card Issuer", 7500, 24.99, 190, 5, "High priority"],
    ],
    "Assets": [
        ["Home", "real_estate", 600000, ""],
        ["Roth IRA", "retirement", 35000, ""],
    ],
    "Budgets": [
        ["2026-06", "Groceries", 800],
        ["2026-06", "Restaurants", 300],
    ],
    "Goals": [
        ["Emergency Fund", 30000, 10000, "2027-12-31", "6 months of expenses"],
        ["Pay Off Truck", 0, 22000, "2027-12-31", ""],
    ],
}


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 35)


def create_template(path: Path = TEMPLATE_PATH) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)

        for row in EXAMPLE_ROWS.get(sheet_name, []):
            ws.append(row)

        ws.freeze_panes = "A2"
        style_header(ws)
        autofit_columns(ws)

    wb.save(path)
    return path


if __name__ == "__main__":
    output_path = create_template()
    print(f"Created template: {output_path}")